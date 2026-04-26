import streamlit as st
import pandas as pd
from io import BytesIO
from PIL import Image
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================
# CONFIGURATION DE LA PAGE
# =========================
st.set_page_config(
    page_title="Packing Dashboard",
    page_icon="📦",
    layout="wide"
)

# =========================
# STYLES CSS PERSONNALISÉS
# =========================
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .metric-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
    }
    .info-text {
        background-color: #e3f2fd;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #2196f3;
    }
    .error-text {
        background-color: #ffebee;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #f44336;
        color: #c62828;
    }
    .download-button {
        margin-top: 50px;
    }
    </style>
""", unsafe_allow_html=True)

# Palette de couleurs pour les modèles
COLOR_PALETTE = [
    '#FFE5B4', '#FFD1DC', '#C4E0FA', '#D4F1F9', '#E6E6FA',
    '#C8E6C9', '#FFCCBC', '#F8BBD9', '#BBDEFB', '#C5E1A5',
    '#FFE0B2', '#D1C4E9', '#B2DFDB', '#F0F4C3', '#FFCDD2'
]

def get_model_color(model_name, model_list):
    """Attribue une couleur unique à chaque modèle"""
    if model_name in model_list:
        index = model_list.index(model_name) % len(COLOR_PALETTE)
        return COLOR_PALETTE[index]
    return COLOR_PALETTE[0]

def style_excel_with_borders_and_colors(file_path, df_result, model_list):
    """Applique des bordures et des couleurs au fichier Excel"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Définir les styles
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
    
    # Appliquer les styles
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style pour l'en-tête
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # Appliquer les couleurs par modèle (colonne A = Model)
    model_colors = {}
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        model_value = row[0].value
        if model_value:
            if model_value not in model_colors:
                model_colors[model_value] = get_model_color(model_value, model_list)
            model_fill = PatternFill(start_color=model_colors[model_value][1:], 
                                    end_color=model_colors[model_value][1:], 
                                    fill_type="solid")
            for cell in row:
                cell.fill = model_fill
    
    wb.save(file_path)

# =========================
# CHARGEMENT DES LOGOS
# =========================
try:
    container_logo = Image.open("conteneur_logo.png")
    stream_logo = Image.open("stream_logo.png")
    logo_container_exists = True
except FileNotFoundError:
    logo_container_exists = False
    st.warning("⚠️ Les fichiers de logos sont introuvables. Vérifiez leur emplacement.")

# =========================
# EN-TÊTE AVEC LOGOS
# =========================
if logo_container_exists:
    col1, col2, col3 = st.columns([1.5, 3, 1.5])
    
    with col1:
        st.image(container_logo, width=120)
    
    with col2:
        st.markdown('<div class="main-header">📊 Packing Summary by Model & Type</div>', 
                   unsafe_allow_html=True)
    
    with col3:
        st.image(stream_logo, width=100)
else:
    st.markdown('<div class="main-header">📊 Packing Summary by Model & Type</div>', 
               unsafe_allow_html=True)

st.markdown("---")

# =========================
# SECTION INFORMATIONS D'EXPÉDITION
# =========================
with st.expander("📋 Informations d'expédition", expanded=True):
    col1, col2 = st.columns(2)
    
    with col1:
        apna = st.text_input(
            "🔖 Numéro APNA",
            placeholder="Ex: APNA-2024-001",
            help="Identifiant unique de l'expédition"
        )
    
    with col2:
        order_shipment = st.text_input(
            "📦 Order of Shipment",
            placeholder="Ex: ORD-001",
            help="Numéro d'ordre d'expédition"
        )

st.markdown("---")

# =========================
# TÉLÉCHARGEMENT DU FICHIER
# =========================
uploaded_file = st.file_uploader(
    "📥 Téléchargez votre fichier Excel",
    type=["xlsx"],
    help="Format accepté: .xlsx"
)

def find_column(df, possible_names):
    """Trouve une colonne dans le DataFrame en essayant plusieurs noms possibles"""
    for name in possible_names:
        for col in df.columns:
            if name.lower() in col.lower():
                return col
    return None

if uploaded_file is not None:
    
    # =========================
    # CHARGEMENT ET NETTOYAGE
    # =========================
    with st.spinner("🔄 Chargement du fichier en cours..."):
        df = pd.read_excel(uploaded_file)
        
        # Nettoyage des noms de colonnes
        original_columns = df.columns.copy()
        df.columns = (
            df.columns
            .astype(str)
            .str.strip()
            .str.replace("\n", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
        )
        
        st.success("✅ Fichier chargé avec succès!")
        
        # Affichage des colonnes trouvées pour débogage
        with st.expander("📋 Colonnes trouvées dans le fichier"):
            st.write("Colonnes disponibles:", list(df.columns))
    
    # =========================
    # DÉTECTION DES COLONNES REQUISES
    # =========================
    # Recherche de la colonne Model
    col_model = find_column(df, ['model', 'modele', 'modèle', 'product', 'article'])
    
    if col_model is None:
        st.markdown("""
        <div class="error-text">
            <strong>❌ Colonne 'Model' introuvable</strong><br>
            Votre fichier doit contenir une colonne pour les modèles (ex: 'Model', 'Modèle', 'Product').<br>
            Colonnes trouvées : {}
        </div>
        """.format(', '.join(list(df.columns))), unsafe_allow_html=True)
        st.stop()
    
    # Recherche de la colonne TYPE
    col_type = find_column(df, ['type', 'categorie', 'category'])
    
    if col_type is None:
        st.markdown("""
        <div class="error-text">
            <strong>❌ Colonne 'TYPE' introuvable</strong><br>
            Votre fichier doit contenir une colonne pour les types (ex: 'Type', 'Catégorie').<br>
            Colonnes trouvées : {}
        </div>
        """.format(', '.join(list(df.columns))), unsafe_allow_html=True)
        st.stop()
    
    # Recherche des colonnes numériques
    col_ctn = find_column(df, ['ctn', 'carton', 'box', 'quantity'])
    col_nw = find_column(df, ['n w', 'net', 'poids net', 'nw', 'weight net'])
    col_gw = find_column(df, ['g w', 'gross', 'poids brut', 'gw', 'weight gross'])
    col_vol = find_column(df, ['volume', 'cbm', 'vol', 'm3'])
    
    missing_cols = []
    if col_ctn is None: missing_cols.append("CTN/Quantité")
    if col_nw is None: missing_cols.append("N W/Poids Net")
    if col_gw is None: missing_cols.append("G W/Poids Brut")
    if col_vol is None: missing_cols.append("Volume/CBM")
    
    if missing_cols:
        st.markdown(f"""
        <div class="error-text">
            <strong>❌ Colonnes manquantes :</strong><br>
            {', '.join(missing_cols)}<br><br>
            <strong>Colonnes trouvées :</strong> {', '.join(list(df.columns))}
        </div>
        """, unsafe_allow_html=True)
        st.stop()
    
    # Conversion en numérique
    for col in [col_ctn, col_nw, col_gw, col_vol]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # =========================
    # FILTRE PAR MODÈLE (SÉLECTION UNIQUE)
    # =========================
    with st.sidebar:
        st.markdown("## 🔧 Filtres")
        models_all = df[col_model].dropna().unique()
        
        if len(models_all) == 0:
            st.error("Aucun modèle trouvé dans les données")
            st.stop()
        
        # Changement : utilisation de selectbox au lieu de multiselect pour n'avoir qu'un seul modèle
        selected_model = st.selectbox(
            "🎯 Sélectionner un modèle",
            options=sorted(models_all),
            help="Sélectionnez un seul modèle à analyser"
        )
    
    df_filtered = df[df[col_model] == selected_model]
    
    # =========================
    # SAISIE DES QUANTITÉS LOT
    # =========================
    st.subheader("📥 Saisie des quantités LOT par type")
    
    types = df_filtered[col_type].unique()
    lot_qty_dict = {}
    
    if len(types) > 0:
        # Création des colonnes dynamiquement
        cols = st.columns(min(len(types), 4))
        
        for idx, type_name in enumerate(types):
            col_idx = idx % 4
            lot_qty_dict[type_name] = cols[col_idx].number_input(
                f"🏷️ {type_name}",
                min_value=0,
                value=0,
                step=100,
                key=f"lot_{type_name}"
            )
    else:
        st.warning("Aucun type disponible après filtrage")
    
    # =========================
    # CALCUL DES AGRÉGATIONS
    # =========================
    with st.spinner("📊 Calcul des statistiques..."):
        result = df_filtered.groupby(
            [col_model, col_type], as_index=False
        ).agg({
            col_ctn: "sum",
            col_nw: "sum",
            col_gw: "sum",
            col_vol: "sum"
        })
        
        result.columns = [
            "Model", "TYPE",
            "CTN QTY",
            "TOTAL N W (KG)",
            "TOTAL G W (KG)",
            "TOTAL VOLUME (CBM)"
        ]
        
        # Ajout des quantités LOT
        result["LOT QTY"] = result["TYPE"].map(lot_qty_dict)
    
    # =========================
    # AFFICHAGE DU TABLEAU STYLISÉ AVEC COULEURS
    # =========================
    st.subheader("📦 Résumé détaillé")
    
    def style_table_with_colors(df):
        """Applique un style professionnel au DataFrame avec couleurs par modèle"""
        # Obtenir la liste des modèles uniques
        models = df['Model'].unique()
        
        # Créer le style de base
        styled = df.style.set_properties(**{
            'border': '1px solid #ddd',
            'text-align': 'center',
            'padding': '8px',
            'font-size': '14px'
        }).set_table_styles([
            {'selector': 'thead tr th',
             'props': [
                 ('background-color', '#1f77b4'),
                 ('color', 'white'),
                 ('border', '1px solid #1f77b4'),
                 ('text-align', 'center'),
                 ('padding', '10px'),
                 ('font-weight', 'bold')
             ]},
            {'selector': 'tbody tr:hover',
             'props': [('background-color', '#f5f5f5')]}
        ])
        
        # Appliquer les couleurs par modèle
        for model in models:
            color = get_model_color(model, models)
            styled = styled.apply(
                lambda x: [f'background-color: {color}' if x['Model'] == model else '' for _ in x],
                axis=1,
                subset=pd.IndexSlice[:, :]
            )
        
        # Formater les nombres
        return styled.format({
            'CTN QTY': '{:,.0f}',
            'TOTAL N W (KG)': '{:,.2f}',
            'TOTAL G W (KG)': '{:,.2f}',
            'TOTAL VOLUME (CBM)': '{:,.3f}',
            'LOT QTY': '{:,.0f}'
        })
    
    st.dataframe(style_table_with_colors(result), use_container_width=True, height=400)
    
    # =========================
    # INDICATEURS CLÉS DE PERFORMANCE
    # =========================
    st.markdown("---")
    st.subheader("📈 Indicateurs globaux")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown("""
        <div class="metric-card">
            <h3>📦</h3>
            <h4>Total CTN</h4>
            <h2>{:,}</h2>
        </div>
        """.format(int(result["CTN QTY"].sum())), unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="metric-card">
            <h3>⚖️</h3>
            <h4>Poids Net</h4>
            <h2>{:,.2f} kg</h2>
        </div>
        """.format(result["TOTAL N W (KG)"].sum()), unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="metric-card">
            <h3>⚖️</h3>
            <h4>Poids Brut</h4>
            <h2>{:,.2f} kg</h2>
        </div>
        """.format(result["TOTAL G W (KG)"].sum()), unsafe_allow_html=True)
    
    with col4:
        st.markdown("""
        <div class="metric-card">
            <h3>📐</h3>
            <h4>Volume Total</h4>
            <h2>{:,.3f} m³</h2>
        </div>
        """.format(result["TOTAL VOLUME (CBM)"].sum()), unsafe_allow_html=True)
    
    # =========================
    # PRÉPARATION DU TÉLÉCHARGEMENT
    # =========================
    safe_apna = apna.strip().replace(" ", "_") if apna else "NO_APNA"
    safe_order = order_shipment.strip().replace(" ", "_") if order_shipment else "NO_ORDER"
    file_name = f"packing_summary_{safe_apna}_{safe_order}.xlsx"
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name="Summary")
        
        # Ajout d'une feuille de métadonnées
        metadata = pd.DataFrame({
            "Information": ["APNA", "Order of Shipment", "Modèle sélectionné", "Date de génération", "Colonnes utilisées"],
            "Valeur": [
                apna if apna else "N/A", 
                order_shipment if order_shipment else "N/A",
                selected_model,
                pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                f"Model: {col_model}, Type: {col_type}, CTN: {col_ctn}, NW: {col_nw}, GW: {col_gw}, Vol: {col_vol}"
            ]
        })
        metadata.to_excel(writer, index=False, sheet_name="Metadata")
        
        # Récupérer le chemin du fichier temporaire pour appliquer les styles
        temp_path = writer.book.filename
        
    # Appliquer les styles avec bordures et couleurs
    if 'temp_path' in locals():
        style_excel_with_borders_and_colors(temp_path, result, result['Model'].unique())
        
        # Lire le fichier modifié
        with open(temp_path, 'rb') as f:
            styled_output = BytesIO(f.read())
    else:
        styled_output = output
    
    # =========================
    # BOUTON DE TÉLÉCHARGEMENT (AVEC ESPACE)
    # =========================
    # Ajout d'espaces pour décaler le bouton vers le bas
    st.markdown("<br><br><br>", unsafe_allow_html=True)
    
    col_download1, col_download2, col_download3 = st.columns([1, 2, 1])
    with col_download2:
        st.download_button(
            label="📥 Télécharger le rapport Excel",
            data=styled_output.getvalue() if 'styled_output' in locals() else output.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
    
    # =========================
    # SECTION INFORMATIONS SUPPLÉMENTAIRES
    # =========================
    with st.expander("ℹ️ Informations sur le rapport"):
        st.markdown(f"""
        <div class="info-text">
            <strong>📋 Détails du traitement :</strong><br>
            - Modèle sélectionné : {selected_model}<br>
            - Types trouvés : {', '.join(map(str, types))}<br>
            - Lignes traitées : {len(df_filtered)}<br>
            - Date de génération : {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
            - Fichier source : {uploaded_file.name}<br>
            - Colonnes détectées :<br>
            &nbsp;&nbsp;&nbsp;• Modèle : {col_model}<br>
            &nbsp;&nbsp;&nbsp;• Type : {col_type}<br>
            &nbsp;&nbsp;&nbsp;• CTN : {col_ctn}<br>
            &nbsp;&nbsp;&nbsp;• Poids Net : {col_nw}<br>
            &nbsp;&nbsp;&nbsp;• Poids Brut : {col_gw}<br>
            &nbsp;&nbsp;&nbsp;• Volume : {col_vol}
        </div>
        """, unsafe_allow_html=True)

else:
    # Affichage d'un message d'accueil lorsqu'aucun fichier n'est chargé
    st.markdown("""
    <div class="info-text" style="text-align: center;">
        <h3>👋 Bienvenue sur le Packing Dashboard</h3>
        <p>Pour commencer, veuillez télécharger votre fichier Excel via le bouton ci-dessus.</p>
        <p>📋 <strong>Format attendu :</strong> Fichier .xlsx contenant des colonnes pour :</p>
        <ul style="text-align: left; display: inline-block;">
            <li>Modèle (ex: "Model", "Modèle", "Product")</li>
            <li>Type (ex: "TYPE", "Type", "Catégorie")</li>
            <li>Quantité CTN (ex: "CTN", "QTY", "Quantity")</li>
            <li>Poids Net (ex: "N W", "Net Weight")</li>
            <li>Poids Brut (ex: "G W", "Gross Weight")</li>
            <li>Volume (ex: "Volume", "CBM")</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
    
    # Affichage d'un exemple
    with st.expander("📖 Voir un exemple du format attendu"):
        example_data = {
            "Model": ["MODEL_A", "MODEL_A", "MODEL_B"],
            "TYPE": ["Type1", "Type2", "Type1"],
            "CTN": [10, 5, 8],
            "N W": [100.5, 50.2, 80.0],
            "G W": [120.0, 60.5, 95.5],
            "VOLUME": [0.5, 0.25, 0.4]
        }
        example_df = pd.DataFrame(example_data)
        st.dataframe(example_df, use_container_width=True)
